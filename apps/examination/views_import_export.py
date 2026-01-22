"""考试题目导入导出视图"""
import os
import pandas as pd
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from apps.users.permissions import IsExamManager
from .models import QuestionBank, Question
from .serializers import QuestionSerializer


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsExamManager])
def import_questions(request):
    """批量导入题目"""
    
    if 'file' not in request.FILES:
        return Response({
            'code': 400,
            'message': '请上传文件'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    file = request.FILES['file']
    file_extension = os.path.splitext(file.name)[1].lower()
    
    # 获取题库ID
    question_bank_id = request.data.get('question_bank_id')
    if not question_bank_id:
        return Response({
            'code': 400,
            'message': '请提供题库ID'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        question_bank = QuestionBank.objects.get(id=question_bank_id)
    except QuestionBank.DoesNotExist:
        return Response({
            'code': 404,
            'message': '题库不存在'
        }, status=status.HTTP_404_NOT_FOUND)
    
    try:
        # 读取Excel文件
        if file_extension in ['.xlsx', '.xls']:
            df = pd.read_excel(file)
        elif file_extension == '.csv':
            df = pd.read_csv(file)
        else:
            return Response({
                'code': 400,
                'message': '不支持的文件格式，请上传Excel或CSV文件'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # 验证必填列
        required_columns = ['题目类型', '题目内容', '正确答案', '分数']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            return Response({
                'code': 400,
                'message': f'缺少必填列: {", ".join(missing_columns)}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # 导入数据
        imported_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                question_type = str(row['题目类型']).strip().lower()
                
                # 处理选项
                options = {}
                if question_type in ['single_choice', 'multiple_choice']:
                    options_str = str(row.get('选项', ''))
                    if options_str:
                        # 解析选项，格式如：A.选项1|B.选项2|C.选项3
                        option_list = []
                        for opt in options_str.split('|'):
                            if '.' in opt:
                                key, value = opt.split('.', 1)
                                option_list.append({'key': key.strip(), 'value': value.strip()})
                        options = {'options': option_list}
                
                # 处理正确答案
                correct_answer = str(row['正确答案']).strip()
                if question_type in ['single_choice', 'judgment']:
                    answer = [correct_answer]
                elif question_type == 'multiple_choice':
                    answer = [opt.strip() for opt in correct_answer.split(',')]
                else:
                    answer = [correct_answer]
                
                # 创建题目
                question_data = {
                    'question_bank': question_bank.id,
                    'question_type': question_type,
                    'title': str(row.get('题目标题', ''))[:100],
                    'content': str(row['题目内容']).strip(),
                    'options': options,
                    'correct_answer': {'answer': answer},
                    'score': float(row['分数']),
                    'difficulty': str(row.get('难度', 'medium')).strip().lower(),
                    'analysis': str(row.get('答案解析', '')),
                    'created_by': request.user.id
                }
                
                serializer = QuestionSerializer(data=question_data)
                if serializer.is_valid():
                    serializer.save()
                    imported_count += 1
                else:
                    errors.append(f"第{index + 2}行: {serializer.errors}")
            
            except Exception as e:
                errors.append(f"第{index + 2}行: {str(e)}")
        
        return Response({
            'code': 200,
            'message': f'导入完成，成功导入 {imported_count} 道题目',
            'data': {
                'imported_count': imported_count,
                'errors': errors[:10]  # 最多显示10个错误
            }
        })
    
    except Exception as e:
        return Response({
            'code': 500,
            'message': f'导入失败: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsExamManager])
def export_questions(request):
    """导出题目数据"""
    
    # 获取筛选参数
    question_bank_id = request.GET.get('question_bank_id')
    question_type = request.GET.get('question_type')
    
    # 构建查询
    queryset = Question.objects.select_related('question_bank', 'created_by').all()
    
    if question_bank_id:
        queryset = queryset.filter(question_bank_id=question_bank_id)
    if question_type:
        queryset = queryset.filter(question_type=question_type)
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "题目数据"
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 定义列
    headers = [
        '题目ID', '题库名称', '题目类型', '题目标题', '题目内容',
        '选项', '正确答案', '分数', '难度', '答案解析', '创建人', '创建时间'
    ]
    
    # 写入表头
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 写入数据
    for row_num, question in enumerate(queryset, 2):
        # 处理选项
        options_str = ''
        if question.options and 'options' in question.options:
            options_list = []
            for opt in question.options['options']:
                options_list.append(f"{opt['key']}.{opt['value']}")
            options_str = '|'.join(options_list)
        
        # 处理正确答案
        correct_answer = ''
        if question.correct_answer and 'answer' in question.correct_answer:
            correct_answer = ', '.join(question.correct_answer['answer'])
        
        ws.cell(row=row_num, column=1, value=question.id)
        ws.cell(row=row_num, column=2, value=question.question_bank.name if question.question_bank else '')
        ws.cell(row=row_num, column=3, value=question.get_question_type_display())
        ws.cell(row=row_num, column=4, value=question.title or '')
        ws.cell(row=row_num, column=5, value=question.content)
        ws.cell(row=row_num, column=6, value=options_str)
        ws.cell(row=row_num, column=7, value=correct_answer)
        ws.cell(row=row_num, column=8, value=question.score)
        ws.cell(row=row_num, column=9, value=question.get_difficulty_display())
        ws.cell(row=row_num, column=10, value=question.analysis or '')
        ws.cell(row=row_num, column=11, value=question.created_by.real_name if question.created_by else '')
        ws.cell(row=row_num, column=12, value=question.created_at.strftime('%Y-%m-%d %H:%M:%S') if question.created_at else '')
        
        # 添加边框
        for col_num in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col_num).border = thin_border
    
    # 调整列宽
    column_widths = [10, 20, 12, 30, 50, 40, 15, 8, 8, 40, 12, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # 创建响应
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="questions_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_question_import_template(request):
    """下载题目导入模板"""
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "题目导入模板"
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 定义列（带*的为必填项）
    headers = [
        '题目类型*', '题目标题', '题目内容*', '选项', '正确答案*',
        '分数*', '难度', '答案解析'
    ]
    
    # 写入表头
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 添加示例数据
    examples = [
        ['single_choice', '设备操作题目1', '设备启动前需要检查什么？', 'A.电源|B.安全装置|C.以上都是', 'C', 2, 'medium', '设备启动前需要全面检查'],
        ['multiple_choice', '安全规范题目', '以下哪些属于安全操作规范？', 'A.戴安全帽|B.穿防护服|C.遵守操作规程|D.全部正确', 'A,B,C,D', 3, 'medium', '所有选项都是正确的安全规范'],
        ['judgment', '判断题示例', '设备运行中可以打开防护罩吗？', '', 'false', 1, 'easy', '设备运行中严禁打开防护罩'],
        ['fill_blank', '填空题示例', '设备操作前必须检查___和___。', '', '电源,安全装置', 2, 'medium', '操作前必须检查电源和安全装置'],
        ['essay', '简答题示例', '请简述设备日常维护的三个要点。', '', '清洁,润滑,检查', 5, 'hard', '日常维护包括清洁、润滑和检查'],
    ]
    
    for row_num, example in enumerate(examples, 2):
        for col_num, value in enumerate(example, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = thin_border
    
    # 添加说明
    ws.cell(row=8, column=1, value='填写说明：')
    ws.cell(row=9, column=1, value='1. 带*号的列为必填项')
    ws.cell(row=10, column=1, value='2. 题目类型可选值: single_choice(单选)、multiple_choice(多选)、judgment(判断)、fill_blank(填空)、essay(简答)')
    ws.cell(row=11, column=1, value='3. 难度可选值: easy(简单)、medium(中等)、hard(困难)')
    ws.cell(row=12, column=1, value='4. 单选题和多选题的选项格式: A.选项1|B.选项2|C.选项3')
    ws.cell(row=13, column=1, value='5. 多选题正确答案多个选项用逗号分隔，如: A,B,C')
    ws.cell(row=14, column=1, value='6. 判断题正确答案填 true 或 false')
    ws.cell(row=15, column=1, value='7. 填空题多个答案用逗号分隔')
    
    # 调整列宽
    column_widths = [15, 20, 40, 40, 15, 8, 8, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # 创建响应
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="question_import_template.xlsx"'
    
    wb.save(response)
    return response