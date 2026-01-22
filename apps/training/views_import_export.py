"""课程导入导出视图"""
import os
import pandas as pd
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from apps.users.permissions import IsTrainingManager
from .models import Course, CourseCategory
from .serializers import CourseSerializer


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsTrainingManager])
def import_courses(request):
    """批量导入课程"""
    
    if 'file' not in request.FILES:
        return Response({
            'code': 400,
            'message': '请上传文件'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    file = request.FILES['file']
    file_extension = os.path.splitext(file.name)[1].lower()
    
    try:
        # 读取Excel文件
        if file_extension in ['.xlsx', '.xls']:
            df = pd.read_excel(file)
        elif file_extension == '.csv':
            df = pd.read_csv(file)
        else:
            return Response({
                'code': 400,
                'message': '不支持的文件格式，请上传Excel或CSV文件'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # 验证必填列
        required_columns = ['课程代码', '课程名称', '课程类型', '时长(分钟)', '学分']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            return Response({
                'code': 400,
                'message': f'缺少必填列: {", ".join(missing_columns)}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # 导入数据
        imported_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                # 获取课程分类
                category_name = row.get('课程分类', '默认分类')
                category, _ = CourseCategory.objects.get_or_create(
                    name=category_name,
                    defaults={'code': category_name[:10].upper()}
                )
                
                # 创建课程
                course_data = {
                    'code': str(row['课程代码']).strip(),
                    'title': str(row['课程名称']).strip(),
                    'description': str(row.get('课程描述', '')),
                    'category': category.id,
                    'course_type': str(row['课程类型']).strip().lower(),
                    'duration': int(row['时长(分钟)']),
                    'credit': float(row['学分']),
                    'instructor': str(row.get('讲师', '')),
                    'passing_score': float(row.get('及格分数', 60)),
                    'status': str(row.get('状态', 'draft')).strip().lower(),
                    'created_by': request.user.id
                }
                
                serializer = CourseSerializer(data=course_data)
                if serializer.is_valid():
                    serializer.save()
                    imported_count += 1
                else:
                    errors.append(f"第{index + 2}行: {serializer.errors}")
            
            except Exception as e:
                errors.append(f"第{index + 2}行: {str(e)}")
        
        return Response({
            'code': 200,
            'message': f'导入完成，成功导入 {imported_count} 条数据',
            'data': {
                'imported_count': imported_count,
                'errors': errors[:10]  # 最多显示10个错误
            }
        })
    
    except Exception as e:
        return Response({
            'code': 500,
            'message': f'导入失败: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsTrainingManager])
def export_courses(request):
    """导出课程数据"""
    
    # 获取筛选参数
    category = request.GET.get('category')
    status_filter = request.GET.get('status')
    
    # 构建查询
    queryset = Course.objects.select_related('category', 'created_by').all()
    
    if category:
        queryset = queryset.filter(category_id=category)
    if status_filter:
        queryset = queryset.filter(status=status_filter)
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "课程数据"
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 定义列
    headers = [
        '课程代码', '课程名称', '课程分类', '课程类型', 
        '时长(分钟)', '学分', '讲师', '及格分数',
        '状态', '报名人数', '完成人数', '创建人', '创建时间'
    ]
    
    # 写入表头
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 写入数据
    for row_num, course in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1, value=course.code)
        ws.cell(row=row_num, column=2, value=course.title)
        ws.cell(row=row_num, column=3, value=course.category.name if course.category else '')
        ws.cell(row=row_num, column=4, value=course.get_course_type_display())
        ws.cell(row=row_num, column=5, value=course.duration)
        ws.cell(row=row_num, column=6, value=course.credit)
        ws.cell(row=row_num, column=7, value=course.instructor or '')
        ws.cell(row=row_num, column=8, value=course.passing_score)
        ws.cell(row=row_num, column=9, value=course.get_status_display())
        ws.cell(row=row_num, column=10, value=course.enrollment_count)
        ws.cell(row=row_num, column=11, value=course.completion_count)
        ws.cell(row=row_num, column=12, value=course.created_by.real_name if course.created_by else '')
        ws.cell(row=row_num, column=13, value=course.created_at.strftime('%Y-%m-%d %H:%M:%S') if course.created_at else '')
        
        # 添加边框
        for col_num in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col_num).border = thin_border
    
    # 调整列宽
    column_widths = [15, 30, 15, 12, 12, 8, 15, 10, 10, 10, 10, 12, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # 创建响应
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="courses_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_course_import_template(request):
    """下载课程导入模板"""
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "课程导入模板"
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 定义列（带*的为必填项）
    headers = [
        '课程代码*', '课程名称*', '课程分类', '课程类型*',
        '时长(分钟)*', '学分*', '课程描述', '讲师',
        '及格分数', '状态', '标签'
    ]
    
    # 写入表头
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 添加示例数据
    example_data = [
        'COURSE001', '设备操作培训', '技术培训', 'online',
        120, 2.0, '生产设备基础操作培训', '张讲师',
        60, 'published', '设备,操作,基础'
    ]
    
    for col_num, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = value
        cell.border = thin_border
    
    # 添加说明
    ws.cell(row=4, column=1, value='填写说明：')
    ws.cell(row=5, column=1, value='1. 带*号的列为必填项')
    ws.cell(row=6, column=1, value='2. 课程类型可选值: online(在线)、offline(线下)、hybrid(混合)')
    ws.cell(row=7, column=1, value='3. 状态可选值: draft(草稿)、published(已发布)、archived(已归档)')
    ws.cell(row=8, column=1, value='4. 时长单位为分钟')
    ws.cell(row=9, column=1, value='5. 及格分数默认为60分')
    
    # 调整列宽
    column_widths = [15, 30, 15, 12, 12, 8, 30, 15, 10, 10, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # 创建响应
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="course_import_template.xlsx"'
    
    wb.save(response)
    return response